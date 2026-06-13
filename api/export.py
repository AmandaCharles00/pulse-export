 import json, io, requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def linregress_slope(xs, ys):
    n = len(xs)
    mx = sum(xs)/n; my = sum(ys)/n
    num = sum((xs[i]-mx)*(ys[i]-my) for i in range(n))
    den = sum((x-mx)**2 for x in xs)
    return num/den if den > 0 else 0

SUPA_URL = 'https://bmeqpzytgedymtkwtnis.supabase.co'
SUPA_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJtZXFwenl0Z2VkeW10a3d0bmlzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Nzc2MTI4OTAsImV4cCI6MjA5MzE4ODg5MH0.z0zXbMZd1zTXQuKYV2-yRhAzFmoqPhGp8r025KCdC5M'

HEADERS = {
    'apikey': SUPA_KEY,
    'Authorization': f'Bearer {SUPA_KEY}',
    'Content-Type': 'application/json',
}

BG='FF0D0F14'; BG2='FF12151C'; BG3='FF181C26'
TEXT='FFCDD6E8'; TEXT2='FF6B7A90'
GREEN='FF00C97A'; GREEN_DIM='FF003D25'
RED='FFFF3D55'; RED_DIM='FF4D0010'
BLUE='FF2D8CFF'; BLUE_DIM='FF0A2A4D'
AMBER='FFF5A623'; WHITE='FFFFFFFF'
HEADER_BG='FF1E2330'

def fill(h): return PatternFill('solid', fgColor=h)
def font(h, bold=False, size=11): return Font(color=h, bold=bold, size=size, name='Arial')
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left(): return Alignment(horizontal='left', vertical='center', wrap_text=True)

def atr_color(val):
    if val is None: return TEXT2, BG3
    if val >= 6:  return 'FF00FF9D', 'FF003D25'
    if val >= 3:  return GREEN,      GREEN_DIM
    if val <= -6: return 'FFFF6B7A', 'FF4D0010'
    if val <= -3: return RED,        RED_DIM
    return TEXT2, BG3

def trend_fmt(trend):
    if 'Escalating' in trend: return GREEN
    if 'Fading' in trend: return RED
    return AMBER

def fetch_all_records():
    MN = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc)
    y, m = now.year, now.month
    
    buckets = []
    # Current month
    from_dt = datetime(y, m, 1, tzinfo=timezone.utc).isoformat()
    buckets.append((from_dt, now.isoformat()))
    # Previous 12 months
    for i in range(1, 13):
        mm = m - i
        yy = y
        while mm <= 0:
            mm += 12; yy -= 1
        d1 = datetime(yy, mm, 1, tzinfo=timezone.utc)
        mm2 = mm + 1
        yy2 = yy
        if mm2 > 12:
            mm2 = 1; yy2 += 1
        d2 = datetime(yy2, mm2, 1, tzinfo=timezone.utc)
        buckets.append((d1.isoformat(), d2.isoformat()))

    all_records = []
    for from_dt, to_dt in buckets:
        url = (f"{SUPA_URL}/rest/v1/extension_history"
               f"?select=ticker,theme,atr_ma,captured_at"
               f"&captured_at=gte.{from_dt}&captured_at=lt.{to_dt}"
               f"&order=captured_at.desc")
        r = requests.get(url, headers={**HEADERS, 'Range': '0-4999', 'Range-Unit': 'items'})
        if r.ok:
            all_records.extend(r.json())
    return all_records

def build_ticker_map(all_records):
    from datetime import datetime
    MN = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    ticker_map = {}
    for rec in all_records:
        ticker = rec['ticker']
        theme = rec['theme']
        atr = float(rec['atr_ma'])
        d = datetime.fromisoformat(rec['captured_at'].replace('Z','+00:00'))
        mk = f"{d.year}-{str(d.month).padStart if False else str(d.month).zfill(2)}"
        mk = f"{d.year}-{str(d.month).zfill(2)}"
        ls = f"{d.month}/{d.day}/{str(d.year)[2:]}"
        if ticker not in ticker_map:
            ticker_map[ticker] = {'theme': theme, 'months': {}}
        ex = ticker_map[ticker]['months'].get(mk)
        if ex is None or abs(atr) > abs(ex['atr']):
            ticker_map[ticker]['months'][mk] = {'ls': ls, 'atr': atr}
    return ticker_map

def compute_trend(atrs):
    n = len(atrs)
    if n < 3: return '→ Stable'
    xs = list(range(n))
    mx = sum(xs)/n; my = sum(atrs)/n
    num = sum((xs[i]-mx)*(atrs[i]-my) for i in range(n))
    den = sum((x-mx)**2 for x in xs)
    slope = num/den if den > 0 else 0
    mean_abs = sum(abs(a) for a in atrs)/n
    norm = slope/mean_abs if mean_abs > 0 else 0
    if norm > 0.08: return '↑ Escalating'
    if norm < -0.08: return '↓ Fading'
    return '→ Stable'

def build_excel(all_records):
    MN = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    
    ticker_map = build_ticker_map(all_records)
    
    month_keys = sorted(set(
        mk for v in ticker_map.values() for mk in v['months']
    ))
    def month_label(mk):
        y, m = mk.split('-')
        return f"{MN[int(m)-1]} {y}"

    # Qualify and analyze
    qualified = []
    for ticker, v in ticker_map.items():
        entries = sorted(v['months'].items())
        if len(entries) < 3: continue
        atrs = [e[1]['atr'] for e in entries]
        above = sum(1 for a in atrs if a > 0)
        below = sum(1 for a in atrs if a < 0)
        total = len(atrs)
        direction = 'Above MA' if above >= below else 'Below MA'
        consistency = round((above if above >= below else below) / total * 100)
        trend = compute_trend(atrs)
        qualified.append({
            'ticker': ticker, 'theme': v['theme'],
            'appearances': total, 'direction': direction,
            'consistency': consistency, 'trend': trend,
            'months': v['months']
        })
    qualified.sort(key=lambda x: x['appearances'], reverse=True)

    escalating = sum(1 for t in qualified if 'Escalating' in t['trend'])
    fading = sum(1 for t in qualified if 'Fading' in t['trend'])
    above_ct = sum(1 for t in qualified if t['direction'] == 'Above MA')
    below_ct = sum(1 for t in qualified if t['direction'] == 'Below MA')

    wb = Workbook()

    # ── ABOUT SHEET ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'About This Report'
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 3
    ws1.column_dimensions['B'].width = 4
    ws1.column_dimensions['C'].width = 54
    for col in range(4, 12):
        ws1.column_dimensions[get_column_letter(col)].width = 14
    for row in range(1, 70):
        for col in range(1, 12):
            ws1.cell(row=row, column=col).fill = fill(BG)

    ws1.row_dimensions[2].height = 40
    ws1.merge_cells('B2:J2')
    c = ws1['B2']
    c.value = 'PULSE SCANNER — Extension Recurrence Report'
    c.font = Font(color=WHITE, bold=True, size=18, name='Arial')
    c.alignment = left(); c.fill = fill(BG)

    ws1.merge_cells('B3:J3')
    c = ws1['B3']
    c.value = 'Tickers appearing 3+ times on the ATR/50MA Extension Scanner'
    c.font = font(TEXT2, size=11); c.alignment = left(); c.fill = fill(BG)

    ws1.row_dimensions[4].height = 5
    for col in range(2, 11):
        ws1.cell(row=4, column=col).fill = fill(BLUE)
    ws1.row_dimensions[5].height = 8

    stats = [
        (f'{len(qualified)} Tickers', 'Appearing 3+ months'),
        (f'{month_label(month_keys[0])} — {month_label(month_keys[-1])}', 'Date Range'),
        (f'{above_ct} Above / {below_ct} Below', 'Direction Split'),
        (f'{escalating} Escalating / {fading} Fading', 'Trend Split'),
    ]
    ws1.row_dimensions[6].height = 28
    ws1.row_dimensions[7].height = 18
    for i, (val, lbl) in enumerate(stats):
        col = 2 + i*2
        ws1.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
        c = ws1.cell(row=6, column=col)
        c.value = val; c.font = Font(color=BLUE, bold=True, size=12, name='Arial')
        c.alignment = center(); c.fill = fill(BG2)
        ws1.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)
        c2 = ws1.cell(row=7, column=col)
        c2.value = lbl; c2.font = font(TEXT2, size=9)
        c2.alignment = center(); c2.fill = fill(BG2)

    def section_header(ws, row, text):
        ws.row_dimensions[row].height = 10
        ws.row_dimensions[row+1].height = 22
        ws.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=10)
        c = ws.cell(row=row+1, column=2)
        c.value = text; c.font = Font(color=AMBER, bold=True, size=10, name='Arial')
        c.alignment = left(); c.fill = fill(BG)
        return row+2

    def bullet(ws, row, text, sub=False):
        ws.row_dimensions[row].height = 20
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        c = ws.cell(row=row, column=2)
        c.value = ('       ' if sub else '▸  ') + text
        c.font = font(TEXT2 if sub else TEXT, size=10)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.fill = fill(BG)
        return row+1

    def use_row(ws, row, num, title, desc):
        ws.row_dimensions[row].height = 18
        ws.row_dimensions[row+1].height = 26
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        c = ws.cell(row=row, column=2)
        c.value = f'  {num}.  {title}'
        c.font = Font(color=GREEN, bold=True, size=10, name='Arial')
        c.alignment = left(); c.fill = fill(BG2)
        ws.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=10)
        c2 = ws.cell(row=row+1, column=2)
        c2.value = '       ' + desc
        c2.font = font(TEXT, size=10)
        c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c2.fill = fill(BG2)
        return row+2

    r = 9
    r = section_header(ws1, r, 'COLUMNS EXPLAINED')
    r = bullet(ws1, r, 'Appearances — Number of distinct months the ticker was 3x+ ATR from its 50MA.')
    r = bullet(ws1, r, 'Direction — Dominant direction: Above MA (bullish extension) or Below MA (bearish/oversold).')
    r = bullet(ws1, r, 'Consistency % — What % of appearances were in the dominant direction. 100% = always same way.')
    r = bullet(ws1, r, 'Trend — Is the extension magnitude growing or shrinking over time?')
    r = bullet(ws1, r, '↑ Escalating = recent months show stronger extensions than earlier months.', sub=True)
    r = bullet(ws1, r, '↓ Fading = extensions getting smaller — possible trend exhaustion or mean reversion.', sub=True)
    r = bullet(ws1, r, '→ Stable = consistent magnitude with no clear directional change.', sub=True)
    r = bullet(ws1, r, 'Peak ATR/50MA — Most extreme reading that month. Positive = above 50MA. Negative = below.')

    r = section_header(ws1, r, 'PRACTICAL TRADING USES')
    uses = [
        ('Mean Reversion Setups', 'If a ticker has gone -5x below its 50MA 8 of 13 months and just hit -4x again, you have historical precedent it recovers. Pattern-match against its own behavior, not intuition.'),
        ('Avoid Fighting the Trend', 'A name extending above its 50MA 11 months in a row is not a short. Knowing its history prevents you from fading a freight train.'),
        ('Sector Rotation Timing', 'When Gold names cluster in the same months, that confirms a macro move. If they cluster again, you have a roadmap for what follows.'),
        ('Position Sizing', 'Names hitting 6-8x ATR extensions regularly have a wider normal range. Size them smaller to keep risk consistent.'),
        ('Watchlist Prioritization', 'You already know which names are historically active — no need to scan 600+ tickers cold every session.'),
        ('Re-Entry Candidates', 'Names that appeared, disappeared 2-3 months, then reappeared are often the best setups — reset near MA then re-extended.'),
        ('Escalating Names to Watch', 'Tickers marked Escalating are extending further over time — momentum is building. High-conviction trend candidates.'),
        ('Fading Names for Exits', 'Tickers marked Fading show weaker extensions recently. If holding based on extension, this signals the move may be exhausting.'),
    ]
    for i, (title, desc) in enumerate(uses):
        r = use_row(ws1, r, i+1, title, desc)

    r = section_header(ws1, r, 'HOW TO READ THE DATA TAB')
    tips = [
        'Sort by Appearances (desc) to find the most chronically active tickers.',
        'Sort by Consistency % to separate pure directional names from oscillators.',
        'Sort by Trend to find Escalating names building momentum right now.',
        'Empty ATR cells = ticker was within 3x ATR of 50MA that month (not extended).',
        'A gap between two populated months = potential re-entry setup after a reset.',
        'Negative ATR values = below 50MA. Positive = above. Bold = 5x+ extension.',
    ]
    for tip in tips:
        r = bullet(ws1, r, tip)

    # ── DATA SHEET ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Extension Recurrence')
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'G3'

    FIXED = 6
    for k, v in {'A':9,'B':26,'C':13,'D':12,'E':14,'F':14}.items():
        ws2.column_dimensions[k].width = v
    for i in range(len(month_keys)):
        ws2.column_dimensions[get_column_letter(FIXED+1+i*2)].width = 11
        ws2.column_dimensions[get_column_letter(FIXED+2+i*2)].width = 12

    total_cols = FIXED + len(month_keys)*2 + 1
    for row in range(1, len(qualified)+10):
        for col in range(1, total_cols):
            ws2.cell(row=row, column=col).fill = fill(BG)

    ws2.row_dimensions[1].height = 22
    ws2.row_dimensions[2].height = 18

    for col, label in enumerate(['TICKER','THEME','APPEARANCES','DIRECTION','CONSISTENCY %','TREND'], 1):
        ws2.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        c = ws2.cell(row=1, column=col)
        c.value = label; c.font = Font(color=TEXT2, bold=True, size=8, name='Arial')
        c.alignment = center(); c.fill = fill(HEADER_BG)

    for i, mk in enumerate(month_keys):
        col_s = FIXED+1+i*2
        ws2.merge_cells(start_row=1, start_column=col_s, end_row=1, end_column=col_s+1)
        c = ws2.cell(row=1, column=col_s)
        c.value = month_label(mk); c.font = Font(color=WHITE, bold=True, size=9, name='Arial')
        c.alignment = center(); c.fill = fill(BLUE_DIM)
        for j, lbl in enumerate(['Last Seen', 'Peak ATR/50MA']):
            c2 = ws2.cell(row=2, column=col_s+j)
            c2.value = lbl; c2.font = Font(color=TEXT2, size=8, name='Arial')
            c2.alignment = center(); c2.fill = fill(HEADER_BG)

    for r_idx, t in enumerate(qualified):
        row = r_idx+3
        ws2.row_dimensions[row].height = 17
        bg = BG if r_idx%2==0 else BG2

        c = ws2.cell(row=row, column=1, value=t['ticker'])
        c.font = Font(color=BLUE, bold=True, size=10, name='Arial')
        c.alignment = left(); c.fill = fill(bg)

        c = ws2.cell(row=row, column=2, value=t['theme'])
        c.font = font(TEXT, size=9); c.alignment = left(); c.fill = fill(bg)

        c = ws2.cell(row=row, column=3, value=t['appearances'])
        c.font = Font(color=AMBER, bold=True, size=10, name='Arial')
        c.alignment = center(); c.fill = fill(bg)

        dir_col = GREEN if t['direction']=='Above MA' else RED
        c = ws2.cell(row=row, column=4, value=t['direction'])
        c.font = Font(color=dir_col, bold=True, size=9, name='Arial')
        c.alignment = center(); c.fill = fill(bg)

        cons = t['consistency']
        cc = GREEN if cons>=80 else AMBER if cons>=60 else TEXT2
        c = ws2.cell(row=row, column=5, value=f"{cons}%")
        c.font = Font(color=cc, bold=True, size=10, name='Arial')
        c.alignment = center(); c.fill = fill(bg)

        tcolor = trend_fmt(t['trend'])
        c = ws2.cell(row=row, column=6, value=t['trend'])
        c.font = Font(color=tcolor, bold=True, size=9, name='Arial')
        c.alignment = center(); c.fill = fill(bg)

        for i, mk in enumerate(month_keys):
            col_s = FIXED+1+i*2
            m = t['months'].get(mk)
            c_ls = ws2.cell(row=row, column=col_s)
            c_ls.fill = fill(bg)
            if m:
                c_ls.value = m['ls']
                c_ls.font = font(TEXT2, size=9)
                c_ls.alignment = center()
            c_atr = ws2.cell(row=row, column=col_s+1)
            if m:
                atr_f = m['atr']
                fc, bgc = atr_color(atr_f)
                c_atr.value = round(atr_f, 2)
                c_atr.font = Font(color=fc, bold=abs(atr_f)>=5, size=10, name='Arial')
                c_atr.alignment = center()
                c_atr.fill = fill(bgc)
                c_atr.number_format = '+0.00;-0.00;0.00'
            else:
                c_atr.fill = fill(bg)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


from http.server import BaseHTTPRequestHandler

class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            all_records = fetch_all_records()
            excel_bytes = build_excel(all_records)
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="extension_recurrence.xlsx"')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Length', str(len(excel_bytes)))
            self.end_headers()
            self.wfile.write(excel_bytes)
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps({'error': str(e), 'traceback': tb}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
